#!/bin/python3

import re 
import os
import sys
import datetime
import openpyxl 
from convert import *
from genMemReg import *
from transRegtxt4Sdk import *

from database import *
sys.path.append('../rawdata/')

def createTblRawData():
    for module in ModuleList:
        tblDict = {}
        with open('../spec/{}.spec'.format(module), 'r') as f:
            lines  = f.readlines()
            regExp = re.compile(r'([DC][st][MRl]\w+)\.(\w+)')
            for line in lines:
                if regExp.findall(line):
                    regList = regExp.findall(line)
                    for i in range(len(regList)):
                        tblName, fldName = regList[i]
                        if re.compile(r'(\d+)?Val\b').search(tblName):
                            tblName = re.sub(r'(\d+)?Val\b', '', tblName)
                        if tblName not in tblDict.keys():
                            tblDict[tblName] = []
                        if fldName not in tblDict[tblName]:
                            tblDict[tblName].append(fldName)

        with open('../rawdata/{}.py'.format(module), 'w') as f:
            f.write('{} = {}'.format(module, str(tblDict)))

def createTblXlsx():
    book   = openpyxl.Workbook()
    red    = openpyxl.styles.Font(color='00FF0000')
    green  = openpyxl.styles.Font(color='00008000')
    tblDir = '../xlsx/tbl.xlsx'

    for moduleName in ModuleList:
        module = book.create_sheet(moduleName)
        module.column_dimensions['A'].width = 25
        module.column_dimensions['B'].width = 20
        module.column_dimensions['C'].width = 10
        module.column_dimensions['D'].width = 10
        module.column_dimensions['E'].width = 15
        module.column_dimensions['F'].width = 20
        module.column_dimensions['G'].width = 10
        offset = 1
        tblDict = getattr(__import__(moduleName), moduleName)
        for key in tblDict.keys():
            module.cell(row=offset, column=1, value='name').font  = red
            module.cell(row=offset, column=2, value='field').font = red
            module.cell(row=offset, column=3, value='width').font = red 
            module.cell(row=offset, column=4, value='opration').font = red 
            module.cell(row=offset, column=5, value='default').font = red 
            module.cell(row=offset, column=6, value='comment').font = red 
            module.cell(row=offset, column=7, value='entries').font = red 
            offset += 1
            for field in tblDict[key]:
                module.cell(row=offset, column=1, value=key).font = green
                module.cell(row=offset, column=2, value=field).font = green
                offset += 1
            offset += 1
        if os.path.exists(tblDir):
            oldBook = openpyxl.load_workbook(tblDir)
            rows = module.max_row
            if module in oldBook.sheetnames:
                oldRows = oldBook[module].max_row
                for i in range(rows):
                    for j in range(oldRows):
                        if module.cell(i+1, 1).value == oldBook[module].cell(j+1, 1).value and module.cell(i+1, 2).value == oldBook[module].cell(j+1, 2).value:
                            module.cell(i+1, 3).value = oldBook[module].cell(j+1, 3).value
                            module.cell(i+1, 4).value = oldBook[module].cell(j+1, 4).value
                            module.cell(i+1, 5).value = oldBook[module].cell(j+1, 5).value
            
    del book['Sheet'] #remove the default sheet before we create, whose name is Sheet
    book.save(tblDir)

def createOrgReg():
    tblDir = '../xlsx/tbl.xlsx'
    regDir = '../orgreg/'
    book   = openpyxl.load_workbook(tblDir)
    for module in ModuleList:
        tblDict = getattr(__import__(module), module)
        if not os.path.exists('../orgreg/'+module):
            os.makedirs('../orgreg/'+module)
            
        rows = book[module].max_row
        for key in tblDict.keys():
            with open('../orgreg/{}/{}.txt'.format(module, key), 'w') as f:
                lines = ''
                title = ''
                for field in tblDict[key]:
                    for i in range(rows):
                        if book[module].cell(i+1, 7).value == 'entries':
                            entries = 1
                            if book[module].cell(i+2, 7).value:
                                entries = book[module].cell(i+2, 7).value
                            title   = 'Name\t{}\tEntryNumber\t{}\n'.format(key, str(entries))
                        if book[module].cell(i+1, 1).value == key and book[module].cell(i+1, 2).value == field:
                            width = 1
                            default = 0
                            opration = 'RO'
                            comment = ''
                            if book[module].cell(i+1, 3).value:
                                width = book[module].cell(i+1, 3).value
                            if book[module].cell(i+1, 4).value:
                                opration = book[module].cell(i+1, 4).value
                            if book[module].cell(i+1, 5).value:
                                default = book[module].cell(i+1, 5).value
                            if book[module].cell(i+1, 6).value:
                                comment = book[module].cell(i+1, 6).value
                            lines  +=  '{}[{}:0]\t{}\t{}\t{}\n'.format(field, str(width-1), opration, str(default), comment)
                f.write(title+lines)

def createPiRawData():
    allPiList = []
    for module in ModuleList:
        piFldList = []
        with open('../spec/{}.spec'.format(module), 'r') as f:
            lines  = f.readlines()
            regExp = re.compile(r'^\s*PI\.(\w+)')
            for line in lines:
                if regExp.search(line):
                    fldName = regExp.search(line).group(1)
                    if fldName not in allPiList:
                        allPiList.append(fldName)
                        piFldList.append(fldName)
                    
        with open('../rawdata/{}PI.py'.format(module), 'w') as f:
            f.write('{} = {}'.format(module, str(piFldList)))

def createPiXlsx():
    book  = openpyxl.Workbook()
    red   = openpyxl.styles.Font(color='00FF0000')
    green = openpyxl.styles.Font(color='00008000')
    piDir = '../xlsx/pi.xlsx'

    for moduleName in ModuleList:
        module = book.create_sheet(moduleName)
        module.column_dimensions['A'].width = 20
        module.column_dimensions['B'].width = 10
        piList = getattr(__import__(moduleName+'PI'), moduleName)
        module.cell(row=1, column=1, value='field').font = red
        module.cell(row=1, column=2, value='width').font = red 
        for i in range(len(piList)):
            module.cell(row=i+2, column=1, value=piList[i]).font = green
        if os.path.exists(piDir):
            oldBook = openpyxl.load_workbook(piDir)
            rows = module.max_row
            if module in oldBook.sheetnames:
                oldRows = oldBook[module].max_row
                for i in range(rows):
                    for j in range(oldRows):
                        if module.cell(i+1, 1).value == oldBook[module].cell(j+1, 1).value:
                            module.cell(i+1, 2).value = oldBook[module].cell(j+1, 2).value
            
    del book['Sheet']
    book.save(piDir)

def piXlsxHdl():
    allConsumDict  = {}
    allProduceDict = {}
    for module in ModuleList:
        book = openpyxl.load_workbook('../xlsx/pi.xlsx')
        rows = book[module].max_row

        allProduceDict[module] = {}
        consumList               = []

        # save produced bus signal for single module
        for i in range(rows):
            width = '1'
            if book[module].cell(i+2, 2).value:
                width = book[module].cell(i+2, 2).value
            if book[module].cell(i+2, 1).value:
                allProduceDict[module][book[module].cell(i+2, 1).value] = width

        # save consumed bus signal
        with open('../spec/{}.spec'.format(module)) as f:
            regExp = re.compile(r'PI\.(\w+)')
            lines = f.readlines()
            for line in lines:
                resultList = regExp.findall(line)
                if resultList:
                    for field in resultList:
                        if field not in allProduceDict[module] and field not in consumList:
                            consumList.append(field)
        allConsumDict[module] = consumList

    # create pi struct for c language
    ppInfo = '/**\n* Packet info datat structures\n*/\n#ifndef __PACKET_INFO__\n#define __PACKET_INFO__\n#endif\n#include <stdint.h>\n\ntypedef struct {\n'
    for module in allProduceDict.keys():
        ppInfo += '//{}:\n'.format(module)
        for field in allProduceDict[module].keys():
            ppInfo += '\tuint32_t\t{} : {};\n'.format(field, allProduceDict[module][field])
    ppInfo += '} tPktInfo;\ntPktInfo PI = {0};'
    with open('../include/ppinfo.h', 'w') as f:
        f.write(ppInfo)

    # add bypass signal
    length = len(ModuleList)
    for module in ModuleList[-1:-1:-1]:
        for consum in allConsumDict[module]:
            while length > 1:
                length -= 1
                if consum not in allProduceDict[ModuleList[length]]:
                    allConsumDict[ModuleList[length]].append(consum)
                else:
                    break

    cheekList = []
    rows = book[ModuleList[0]].max_row
    if allConsumDict[ModuleList[0]]:
        print('you need define: \033[1;35m pi.{} \033[0m'.format(str(allConsumDict[ModuleList[0]])))
    for i in range(rows):
        cheekList.append(book[ModuleList[0]].cell(i+1, 1).value)
    for field in allConsumDict[ModuleList[1]]:
        if field not in cheekList:
            print('you need define: \033[1;35m pi.{} \033[0m'.format(field))
    
    # create file to save 2 .vh file and 1 .h file for verilog
    # you need to change the following list when you adjust modules
    #transSiglList = ['IgrHa2PaPi', 'IgrPa2VlPi', 'IgrVl2IfPi', 'IgrIf2LmPi', 'IgrLm2PpPi', 'IgrPp2PoPi', 'IgrPo2RcPi', 'IgrRc2FwPi', 'IgrFw2TmPi', 'EgrTm2EgPi']
    #prefixList    = ['IGR_HA2_PA_', 'IGR_PA2_VL_', 'IGR_VL2_IF_', 'IGR_IF2_LM_', 'IGR_LM2_PP_']
    #for index in range(0, len(ModuleList)-1, 1):
    #    print(ModuleList[index+1])
    #    rows = book[ModuleList[index+1]].max_row
    #    for i in range(rows):
    #        print(book[ModuleList[index+1]].cell(i+1, 1).value)

    #    with open('../bus/{}.h'.format(), 'w') as f:
                  
#def createSpec4vhdl():
    specDir = '../spec/'
    hdlDir  = '../spec4verilog/'
    tblDir  = '../xlsx/tbl.xlsx'
    book    = openpyxl.load_workbook(tblDir)
    for module in ModuleList:
        rows = book[module].max_row  
        newLines = ''
        with open(specDir+module+'.spec', 'r') as f:
            regExp = re.compile(r'(([DC][st][MRl]\w+)\.(\w+))')
            lines = f.readlines()
            for line in lines:
                resultList = regExp.findall(line)
                if resultList and '[' not in line:
                    for i in range(len(resultList)):
                        for j in range(rows):
                            if resultList[i][1] == book[module].cell(j+1, 1).value and resultList[i][2] == book[module].cell(j+1, 2).value:
                                width = '0'
                                if book[module].cell(j+1, 3).value:
                                    width = str(book[module].cell(j+1, 3).value)
                                line = line.replace(resultList[i][0], resultList[i][0]+'[{}:0]'.format(width))
                newLines += line

        with open(hdlDir+module+'.spec', 'w') as f:
            f.write(newLines)

def main():
    #createTblRawData()
    #createTblXlsx()
    #createOrgReg()
    #createRegTxt('../orgreg/', '../regtxt/', '../bak/', '../bak/')
    transRegTxt2C(os.getcwd(), '../regtxt/', '../../include/mem/', '../../lib/mem/')
    #
    #createPiRawData()
    #createPiXlsx()
    #piXlsxHdl()

    funcExtern = '#ifndef __CMODEL_H__\n#define __CMODEL_H__\n\n'
    for module in ModuleList:
        funcExtern = transSpec2Cmodel('../spec/', '../cmodel/', module+'.spec', funcExtern)
    funcExtern += '#endif'
    with open('../../include/cmodel.h', 'w') as f:
        f.write('/*\n * cmodel for Pishon\n * @file: cmodel.h\n * @description: auto generage\n * @note: Copyright (C) {}\n *\n * You should have received a copy of the Fisilink License\n * along with this program; if not, write to Fisilink\n */\n\n'.format(datetime.datetime.now()))
        f.write(funcExtern)
    #createSpec4vhdl()


if __name__ == '__main__':
    main()
